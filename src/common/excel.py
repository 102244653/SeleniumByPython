from src.common.log import Logger
from openpyxl import Workbook, load_workbook
log = Logger().get_logger()


class MyExcel(object):

    def __init__(self, path=None):
        if path is None:
            self.wb = Workbook()
        else:
            self.wb = load_workbook(path)

    def creat_sheet(self, sheet_name, index=0):
        """
        创建并返回表sheet
        :param sheet_name:
        :param index:表的索引位置
        :return:
        """
        return self.wb.create_sheet(sheet_name, index)

    def creat_cell(self, sheet_name,  row=10, column=10):
        """
        创建单元格
        :param row:
        :param column:
        :param sheet_name:
        :return:
        """
        for i in range(1, row):
            for j in range(1, column):
                self.get_sheet(sheet_name).cell(row=i, culumn=j)

    def get_sheet(self, sheet_name):
        """
        得到一个表对象
        :param sheet_name:
        :return:
        """
        return self.wb.get_sheet_by_name(sheet_name)

    def set_value_by_point(self, sheet_name, point, value):
        """
        数据可以直接分配到单元格中(可以输入公式)，例如：A1/B3
        :param sheet_name:
        :param point:
        :param value:
        :return:
        """
        ws = self.get_sheet(sheet_name)
        ws[point] = value

    def set_value_by_row(self, sheet_name, value, row=None):
        """
        可以附加行，从第一列开始附加(从最下方空白处，最左开始)(可以输入多行)
        :param row: 起始行
        :param value:
        :return:
        """
        ws = self.get_sheet(sheet_name)
        if row is None:
            ws.append(value)




